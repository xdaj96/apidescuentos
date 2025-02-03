from DAO.productos_dao import ProductoDAO 
from flask import request,Blueprint, send_file,jsonify
from utils.response import apiresponse
from typing import List,Optional
from libs.JWT import token_requerido
from flask_cors import cross_origin 
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill, Font,Border,borders,Alignment,Side
from io import BytesIO
 
 
DAOProductos = ProductoDAO()
productos_controller = Blueprint('productos',__name__)

@productos_controller.route('/api/productos',methods=['GET'])
@token_requerido
def listar_productos(current_user):
    """ 
    Retorna la lista de los productos 
        
    Returns:
        List[dict]: datos de los productos 
    """
    # definimos los parametros de consulta 
    req_data = request.args.to_dict()
    filtros = {}
    query_value = "" 

    # Si el usuario usa el 'search'
    if 'q' in req_data:
        query_value = req_data['q']
         
        # Verificar si el valor es un número con 13 dígitos (posible código de barras)
        
        if query_value.isdigit() and len(query_value) == 13:
            filtros['cod_barraspri'] = {"valor": query_value , "tipo": "="}
           
        # Verificar si el valor es un número con entre 4 y 12 dígitos (posible número de troquel)
        elif query_value.isdigit() and 4 <= len(query_value) < 13:
            filtros['nro_troquel'] = {"valor": query_value, "tipo": "="}

        # Si no es un número, es una búsqueda general
        else:
            filtros['search'] = {"valor": query_value, "tipo": "LIKE"}
 

        # Ejecutar la consulta con los filtros aplicados
    
    lista_descuentos = DAOProductos.get_productos_paginados(filtros= filtros)
    return  apiresponse(True,data = lista_descuentos)

@productos_controller.route('/api/productos/excel',methods=['GET'])
def download_excel():
        # Crear un nuevo libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"

        # Estilo para las cabeceras (con fondo de color y texto en negrita)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo
        header_font = Font(bold=True)  # Texto en negrita

        # Estilo para las filas alternadas
        row_fill_odd = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris claro
        row_fill_even = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Azul claro

        # Agregar las cabeceras
        ws['A1'] = "Cod. de Barras"
        ws['B1'] = "Nombre de producto"
        ws['C1'] = "Repeticion"
        ws['D1'] = "Unidades"

        # Aplicar estilo de cabecera
        ws['A1'].fill = header_fill
        ws['B1'].fill = header_fill
        ws['C1'].fill = header_fill
        ws['D1'].fill = header_fill
        ws['A1'].font = header_font
        ws['B1'].font = header_font
        ws['C1'].font = header_font
        ws['D1'].font = header_font
        
         # Estilos de alineación
        center_alignment = Alignment(horizontal="center", vertical="center")  # Alineación centrada

        # Estilo para los bordes
        border_style = Border(
            left=Side(border_style="thin", color="000000"),  # Borde izquierdo
            right=Side(border_style="thin", color="000000"),  # Borde derecho
            top=Side(border_style="thin", color="000000"),  # Borde superior
            bottom=Side(border_style="thin", color="000000")  # Borde inferior
        )
    
        
        # Agregar algunas filas de ejemplo
        data = [
            [44322, "Carolina Herrera", '',5],
            [44322, "Carolina Herrera", '',5],
            [44322, "Carolina Herrera", '',5],
            [44322, "Carolina Herrera", '',5],
            [],
            [],[],[],[],[],[],[],[],[],[],[]
        ]
        
            # Cambiar el ancho de las columnas
        ws.column_dimensions['A'].width = 20 # Columna A: Ancho 5
        ws.column_dimensions['B'].width = 40  # Columna B: Ancho 20
        ws.column_dimensions['C'].width = 15  # Columna C: Ancho 10
        ws.column_dimensions['D'].width = 15  # Columna D: Ancho 10

        # Cambiar el alto de las filas
        ws.row_dimensions[1].height = 25  # Fila 1 (cabecera): Alto 25
        ws.row_dimensions[2].height = 20  # Fila 2 (primer dato): Alto 20
        ws.row_dimensions[3].height = 20  # Fila 3: Alto 20
        ws.row_dimensions[4].height = 20  # Fila 4: Alto 20
        ws.row_dimensions[5].height = 20  # Fila 5: Alto 20
    

        # Insertar los datos con color alternado en las filas
        for idx, row in enumerate(data, start=2):  # Comienza en la fila 2 porque la fila 1 es la cabecera
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=idx, column=col, value=value)

                # Aplicar la alineación centrada
                cell.alignment = center_alignment

               
                # Aplicar borde alrededor de cada celda
                cell.border = border_style
        # Guardar el archivo Excel en memoria utilizando BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Volver al inicio del archivo en memoria

        # Enviar el archivo como respuesta para la descarga
        return send_file(output, as_attachment=True, download_name="importar_productos.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
 
    
@productos_controller.route('/api/productos/importar', methods=['POST'])
def importar_productos():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400

    file = request.files['file']
    
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    if file and file.filename.endswith('.xlsx'):
        # Leemos el archivo Excel en memoria
        in_memory_file = BytesIO(file.read())
        workbook = load_workbook(in_memory_file)
        sheet = workbook.active  # Usamos la hoja activa del archivo

        # Extraer los datos de la hoja de cálculo
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Comienza desde la fila 2 para omitir la cabecera
            data.append({
                "cod_barraspri": row[0],
                "nom_producto": row[1],
                "repeticion": row[2],
                "unidades": row[3]

            })

        # Devolver los datos leídos en formato JSON
        return apiresponse(True,'Los datos se importaron correctamente',data,200)

    return jsonify({"error": "Invalid file format. Please upload an .xlsx file."}), 400
